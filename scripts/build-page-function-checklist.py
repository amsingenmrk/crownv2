#!/usr/bin/env python3
"""Simple page-by-page FUNCTIONAL checklist (does this function work on this page?).
Curated from the full component/page catalog; frontend/UI-chrome items intentionally dropped."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = "/Users/as150464/Documents/Vibes V3/glassbox/Glassbox-Page-Function-Checklist.xlsx"

# (Page label, route, [ (function, what it does) , ... ])
PAGES = [
("Global (all pages)", "sidebar / topbar / palette", [
    ("Sidebar navigation", "Links to Portfolio, assets, scenarios, Compare, Benchmarks, Search, Documents all navigate"),
    ("Command palette", "Cmd/Ctrl-K opens; searching recent assets / pages / scenarios / assets navigates"),
    ("Create new scenario", "'+' in sidebar creates a scenario and routes to it"),
    ("Create new portfolio group / fund", "'+' creates a portfolio scope/group"),
    ("Rename scenario", "Rename from breadcrumb/sidebar persists everywhere"),
    ("Delete scenario", "Delete removes it and clears its stored state"),
    ("Rename / delete portfolio group", "Group rename/description/delete persists"),
    ("Recent assets", "Visiting an asset adds it to the recent list in the palette"),
    ("Theme toggle", "Light/dark toggle (and 'D' hotkey) persists across reloads"),
    ("Breadcrumb", "Topbar breadcrumb reflects the current asset/scenario/scope"),
]),
("Login", "/login, /login2", [
    ("Sign in", "Submitting the email form routes to /portfolio (NOTE: no real auth)"),
    ("Contact-for-pricing dialog", "Pricing/contact link opens its dialog"),
]),
("Landing", "/landing", [
    ("Marketing content renders", "Hero, feature grid, footer display"),
    ("Log-in link", "Navigates to the sign-in page"),
    ("Contact form", "Contact-for-pricing form submits / opens"),
]),
("Portfolio overview", "/portfolio", [
    ("Asset table lists portfolio", "All owned assets shown with financial metrics"),
    ("Sort columns", "Column sorting works and persists"),
    ("Show/hide & reorder columns", "Column visibility/order via view options persists"),
    ("Search / filter assets", "Text filter narrows the table"),
    ("Table ↔ map toggle", "Switching between table and map views works"),
    ("Map shows asset pins", "Pins render (color-coded by lift); fallback layout when Mapbox off"),
    ("Map pin → asset", "Clicking a pin / summary card opens the asset"),
    ("KPI strip aggregates", "Occupancy, NOI, value, cap rate roll up correctly"),
    ("Valuation condition toggle", "In-Place / Mark-to-Market / Gross Potential updates KPIs"),
    ("Row → asset detail", "Clicking a row navigates to the property"),
    ("Overview ↔ Forecasts tabs", "Tab switch works"),
]),
("Portfolio forecasts", "/portfolio/forecasts", [
    ("Forecast rolls up all assets", "Aggregated statement/projection for the whole portfolio"),
    ("Outlook weighting", "Baseline/optimistic/pessimistic probability slider changes the expected model"),
    ("Chart metric tabs", "Revenue / OpEx / NOI / Value / Cap Rate tabs switch series"),
    ("Uncertainty bands", "Fan/uncertainty shading renders"),
    ("KPI summary + deltas", "Summary KPIs with base→scenario deltas"),
    ("Per-asset selection in table", "Choosing per-asset versions/outlooks feeds the rollup"),
]),
("Portfolio fund/scope", "/portfolio/scopes/[scopeId] (+ /forecasts)", [
    ("Scoped to fund", "Table/map/KPIs filtered to the selected fund (Fund I/II/III)"),
    ("Scope overview + forecasts", "Both surfaces work scoped to the fund"),
    ("Canonical slug redirect", "Legacy/old scope URLs redirect to the canonical slug"),
]),
("Asset · Stacking Plan", "/properties/[id]/stacking-plan", [
    ("Stacking matrix renders", "Floor / suite / tenant rent roll for the asset"),
    ("Viz mode toggle", "Detailed vs simplified view switches"),
    ("Suite tooltip / drawer", "Hover tooltip and suite detail drawer open"),
    ("Edit leasing assumptions", "Per-suite/building leasing assumption edits persist"),
    ("Split / merge vacant suite", "Vacant space split/merge works"),
    ("Export", "CSV/export of the stacking plan"),
    ("Value-driver waterfall", "Per-floor rent value-driver waterfall renders"),
    ("Occupancy / KPI strip", "Occupancy bar and asset KPIs correct"),
    ("Asset tab navigation", "Stacking/Modifications/Forecasts/Benchmarks tabs switch"),
]),
("Asset · Modifications", "/properties/[id]/modifications", [
    ("Toggle modifications", "Gym / Bar / Cafe / Restaurant / LEED options select"),
    ("Impact on stacking viz", "Rent-lift impact colors update per suite (incl. negative F&B drag)"),
    ("Save modification set", "Create + name a saved set (persists)"),
    ("Apply / rename / delete set", "Manage saved modification sets"),
    ("Draft vs applied state", "Draft changes vs applied set behave correctly"),
    ("Impact filters / legend", "Filter/legend controls work"),
    ("Valuation KPI deltas", "KPI strip shows baseline-vs-modified deltas"),
    ("Recommended modification", "Recommended-mod CTA computes and applies"),
]),
("Asset · Forecasts", "/properties/[id]/forecasts", [
    ("Forecast statement renders", "8-quarter pro-forma for the asset"),
    ("Economic outlook selector", "Pick preset or edit custom macro periods"),
    ("Building-version selector", "Choose baseline or a saved modification set"),
    ("Save / select outlook sets", "Named outlook sets save and load"),
    ("Statement + space rows", "Statement table with expandable space-level rows"),
    ("Chart tabs + uncertainty", "Metric tabs and uncertainty bands"),
    ("KPI summary + deltas", "Summary KPIs with deltas"),
    ("Leasing assumptions editing", "Assumptions edits feed the model"),
]),
("Asset · Benchmarks", "/properties/[id]/benchmarks", [
    ("Resolves asset's market", "Asset auto-mapped to its market & submarket"),
    ("3-column comparison", "Asset vs two benchmark areas"),
    ("Change benchmark area", "Per-column searchable area picker"),
    ("KPI comparison", "KPIs grouped by section across the 3 columns"),
    ("Percentile badges", "Percentile-vs-area badges compute"),
    ("Projection charts", "Intrinsic rent / cap rate projection lines"),
    ("Header map → explorer", "Column map links to /benchmarks?area="),
]),
("Scenario overview", "/scenarios/[slug], /scenarios/2026-capital-planning", [
    ("Scenario scopes assets", "Shows only the scenario's asset set"),
    ("Include / exclude assets", "Add or remove assets from the scenario"),
    ("Add asset from portfolio/search", "Adding an asset elsewhere lands it in the scenario"),
    ("Per-asset modification selection", "Choose a modification set per asset"),
    ("Per-asset outlook selection", "Choose an economic outlook per asset"),
    ("Scenario KPI rollup + deltas", "Scenario-level KPIs with deltas vs baseline"),
    ("Overview ↔ Forecasts tabs", "Tab switch works"),
]),
("Scenario forecasts", "/scenarios/[slug]/forecasts", [
    ("Aggregated scenario forecast", "Rolls up all member assets"),
    ("Selections feed rollup", "Per-asset building-version + outlook choices apply"),
    ("KPI deltas vs baseline", "Deltas reflect the scenario's modifications/outlooks"),
    ("Charts / metric tabs", "Projection charts switch metrics"),
]),
("Compare", "/compare, /compare/new, /compare/[id]", [
    ("Add / remove columns", "Add comparison columns: portfolio / group / scenario / asset"),
    ("Comparison table builds", "Side-by-side KPIs per column"),
    ("Comparison charts", "Forecast charts per column"),
    ("Save comparison", "Name and save the comparison"),
    ("Load saved comparison", "Open a saved comparison by id"),
    ("Rename / delete saved", "Manage saved comparisons"),
]),
("Benchmarks explorer", "/benchmarks", [
    ("Area search", "Search markets / submarkets / counties / ZIPs (geocode + hierarchy)"),
    ("Drilldown breadcrumb", "Nation → market → submarket → county → ZIP navigation"),
    ("?area URL sync", "Opening /benchmarks?area= focuses that area"),
    ("Map boundary renders", "Selected area boundary draws on the map"),
    ("Child areas clickable", "Clicking a child area on the map drills down"),
    ("Area KPI snapshot", "Stats panel shows the area's KPIs + methodology"),
    ("Compare-to-asset", "Pick an in-scope asset → opens its benchmarks"),
    ("Market forecast section", "Editable outlook scenarios + save/load outlook sets"),
]),
("Search", "/search", [
    ("Map of portfolio + market", "Both portfolio assets and market listings on the map"),
    ("Text search", "Search by address/building"),
    ("Scope filter", "Filter portfolio vs market (draft/apply/clear)"),
    ("Listing cards", "Cards show RSF/Occ/Value/Cap + lift"),
    ("Add to group / scenario", "'+' adds a listing to a portfolio group or scenario"),
    ("Card → asset detail", "Card links to the property"),
]),
("Documents", "/documents", [
    ("Upload categories shown", "Rent Roll / T12 / Offering Memorandum cards render"),
    ("File upload", "PLACEHOLDER — 'Select file' is not wired; parsing not built"),
]),
]

# ---- styling ----
NAVY="1F2937"; WHITE="FFFFFF"; BAND="EEF2FF"; GREY="6B7280"
thin=Side(style="thin",color="D1D5DB"); border=Border(left=thin,right=thin,top=thin,bottom=thin)
HF=Font(bold=True,color=WHITE,size=11); CF=Font(size=10)
PAGEF=Font(bold=True,size=10,color="111827")
WRAP=Alignment(wrap_text=True,vertical="top"); CTR=Alignment(horizontal="center",vertical="center")
WORKS='"Yes,No,Partial,Not built,N/A"'

wb=Workbook()

# Read Me
ws=wb.active; ws.title="Read Me"; ws.sheet_view.showGridLines=False
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=105
def line(t,bold=False,size=11,color="111827"):
    ws.append([None,t]); c=ws.cell(ws.max_row,2)
    c.font=Font(bold=bold,size=size,color=color); c.alignment=Alignment(wrap_text=True,vertical="top")
line("Glassbox — Page Function Checklist",bold=True,size=18)
line("For each page: does this function work? One row per real capability — UI/styling detail left out.",color=GREY)
line("")
line("How to use",bold=True,size=13)
line("• Go page by page. For each function, set Works? = Yes / No / Partial / Not built / N/A.")
line("• Use the Notes column for what's broken or missing.")
line("• Filter the 'Works?' column to find everything not yet Yes.")
line("")
line("Note: there's no backend — all data is synthetic and saved in the browser, so 'works' = the action happens and persists on reload.",color=GREY)

# Checklist sheet
ws=wb.create_sheet("Page Functions")
headers=["#","Page","Route","Function","What it does","Works?","Notes"]
widths=[4,26,30,32,62,12,40]
ws.append(headers)
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
n=0; band=False; rowi=1
for label,route,funcs in PAGES:
    band=not band
    for j,(fn,detail) in enumerate(funcs):
        n+=1; rowi+=1
        ws.append([n, label if j==0 else "", route if j==0 else "", fn, detail, "", ""])
        fill=PatternFill("solid",fgColor=BAND) if band else None
        for c in range(1,8):
            cell=ws.cell(rowi,c)
            cell.font=PAGEF if (c in (2,3) and j==0) else CF
            cell.alignment=CTR if c==6 else WRAP
            cell.border=border
            if fill: cell.fill=fill
# header style
for c in range(1,8):
    cell=ws.cell(1,c); cell.font=HF; cell.fill=PatternFill("solid",fgColor=NAVY)
    cell.alignment=Alignment(vertical="center",wrap_text=True); cell.border=border
ws.row_dimensions[1].height=26
ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:G1"
dv=DataValidation(type="list",formula1=WORKS,allow_blank=True)
dv.add(f"F2:F{rowi}"); ws.add_data_validation(dv)

wb.save(OUT)
print("Wrote",OUT,"|",n,"functions across",len(PAGES),"pages")

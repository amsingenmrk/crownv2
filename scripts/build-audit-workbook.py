#!/usr/bin/env python3
"""Build the Glassbox functionality-audit workbook from the catalog workflow JSON."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SRC = "/private/tmp/claude-503/-Users-as150464-Documents-Vibes-V3-glassbox/08ca5243-aedd-47b7-becd-e6ba98783c8f/tasks/wxkck8opp.output"
OUT = "/Users/as150464/Documents/Vibes V3/glassbox/Glassbox-Functionality-Audit.xlsx"

data = json.load(open(SRC))["result"]
pages, components, checklist = data["pages"], data["components"], data["checklist"]

# ---- styling tokens ----
NAVY = "1F2937"        # header bg
NAVY_FONT = "FFFFFF"
BAND = "F3F4F6"        # zebra band
TITLE = "111827"
SECTION_FILL = "E5E7EB"
thin = Side(style="thin", color="D1D5DB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
HEADER_FONT = Font(bold=True, color=NAVY_FONT, size=11, name="Calibri")
CELL_FONT = Font(size=10, name="Calibri")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

PAGE_AREA_ORDER = ["Auth/Landing", "Portfolio", "Asset Detail", "Scenarios", "Compare", "Benchmarks", "Other"]
COMP_CAT_ORDER = ["Shell/Nav", "Portfolio", "Asset Detail", "Stacking Plan", "Modifications",
                  "Forecasts", "Scenarios", "Compare", "Benchmarks", "Landing/Auth", "Shared/Util"]
CHECK_AREA_ORDER = ["Global shell / navigation", "Portfolio dashboard", "Stacking Plan", "Modifications",
                    "Asset Forecasts", "Portfolio/Scenario Forecasts", "Scenarios", "Compare",
                    "Benchmarks", "Asset Benchmarks"]

STATUS_OPTS = '"Not started,In progress,Done,Needs work,Placeholder,N/A"'
WORKS_OPTS = '"Pass,Fail,Partial,Not built,N/A"'

def order_by(rows, key, order):
    idx = {v: i for i, v in enumerate(order)}
    return sorted(enumerate(rows), key=lambda t: (idx.get(t[1].get(key), 999), t[0]))

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

def write_sheet(ws, headers, rows, widths, status_cols=None, band_key_col=2):
    """status_cols: dict {col_index(1-based): formula} for dropdowns."""
    ws.append(headers)
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    prev_band = None
    band_on = False
    for r_i, row in enumerate(rows, start=2):
        ws.append(row)
        # zebra by the grouping column value
        band_val = row[band_key_col - 1]
        if band_val != prev_band:
            band_on = not band_on
            prev_band = band_val
        fill = PatternFill("solid", fgColor=BAND) if band_on else None
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r_i, column=c)
            cell.font = CELL_FONT
            cell.alignment = CENTER if (status_cols and c in status_cols) else WRAP_TOP
            cell.border = border
            if fill:
                cell.fill = fill
    style_header(ws, len(headers))
    if status_cols:
        for col, formula in status_cols.items():
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.add(f"{get_column_letter(col)}2:{get_column_letter(col)}{len(rows) + 1}")
            ws.add_data_validation(dv)

wb = Workbook()

# ================= READ ME =================
ws = wb.active
ws.title = "Read Me"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 110
def line(txt, bold=False, size=11, color="111827", top=0):
    ws.append([None, txt])
    c = ws.cell(row=ws.max_row, column=2)
    c.font = Font(bold=bold, size=size, color=color, name="Calibri")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if top:
        ws.row_dimensions[ws.max_row].height = top
line("Glassbox — Functionality Audit", bold=True, size=18)
line("Page, component, and feature inventory for verifying that all functionality is built and working.", size=11, color="6B7280")
line("")
line("How to use this workbook", bold=True, size=13)
line("1.  Pages  — every route in the app (30). One row per page. Read 'Sub-Features to Verify' and check each works.")
line("2.  Components  — every feature component (89), grouped by area. 'Used By' shows where each is rendered.")
line("3.  Functionality Checklist  — 331 discrete, testable behaviors with acceptance criteria. This is the main walkthrough list.")
line("")
line("Fill in as you review:", bold=True, size=12)
line("•  Status / Works?  — pick from the dropdown in each row (Pages & Components: Status; Checklist: Works?).")
line("•  Reviewed By / Notes/Gaps  — who checked it and anything missing or broken.")
line("•  Use the column filters (row 1) to focus on one area, or sort by Status to find gaps.")
line("")
line("Status values:  Not started · In progress · Done · Needs work · Placeholder · N/A", color="6B7280")
line("Works? values:  Pass · Fail · Partial · Not built · N/A", color="6B7280")
line("")
line("Context", bold=True, size=12)
line("Glassbox is a CRE portfolio analytics demo (Next.js 16 / React 19). No backend: all data is synthetic & "
     "deterministic; all user edits persist to localStorage. Only network calls are to Mapbox. So 'functionality' = "
     "the client-side interactions, persistence, and rendering listed here.")
line("")
line(f"Inventory: {len(pages)} pages · {len(components)} components · {len(checklist)} checklist items.", color="6B7280")
line("Generated from a full read of the codebase (every page + component file).", color="6B7280")

# ================= PAGES =================
ws = wb.create_sheet("Pages")
headers = ["#", "Area", "Route", "Page File", "Layout", "Purpose", "Key Components",
           "How It Works", "Data Source", "Persistence", "Sub-Features to Verify",
           "Status", "Reviewed By", "Notes / Gaps"]
widths = [4, 14, 26, 30, 24, 38, 30, 46, 22, 26, 50, 14, 14, 32]
rows = []
for n, (_, p) in enumerate(order_by(pages, "area", PAGE_AREA_ORDER), 1):
    rows.append([n, p.get("area",""), p.get("route",""), p.get("pageFile",""), p.get("layoutFile",""),
                 p.get("purpose",""), p.get("keyComponents",""), p.get("howItWorks",""),
                 p.get("dataSource",""), p.get("persistence",""), p.get("subFeatures",""),
                 "", "", ""])
write_sheet(ws, headers, rows, widths, status_cols={12: STATUS_OPTS}, band_key_col=2)

# ================= COMPONENTS =================
ws = wb.create_sheet("Components")
headers = ["#", "Category", "Component", "File", "Type", "Description", "Used By",
           "Sub-Features to Verify", "State / Persistence", "Viz",
           "Status", "Reviewed By", "Notes / Gaps"]
widths = [4, 16, 30, 34, 16, 44, 34, 50, 30, 14, 14, 14, 32]
rows = []
for n, (_, c) in enumerate(order_by(components, "category", COMP_CAT_ORDER), 1):
    rows.append([n, c.get("category",""), c.get("name",""), c.get("file",""), c.get("type",""),
                 c.get("description",""), c.get("usedBy",""), c.get("subFeatures",""),
                 c.get("statePersistence",""), c.get("viz",""), "", "", ""])
write_sheet(ws, headers, rows, widths, status_cols={11: STATUS_OPTS}, band_key_col=2)

# ================= CHECKLIST =================
ws = wb.create_sheet("Functionality Checklist")
headers = ["#", "Area", "Feature", "Expected Behavior (Acceptance Criteria)",
           "Where (route / component)", "Works?", "Notes / Gaps"]
widths = [4, 26, 40, 78, 42, 12, 36]
rows = []
for n, (_, c) in enumerate(order_by(checklist, "area", CHECK_AREA_ORDER), 1):
    rows.append([n, c.get("area",""), c.get("feature",""), c.get("detail",""),
                 c.get("relatedTo",""), "", ""])
write_sheet(ws, headers, rows, widths, status_cols={6: WORKS_OPTS}, band_key_col=2)

# ================= SUMMARY =================
ws = wb.create_sheet("Summary")
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 34
ws.column_dimensions["C"].width = 12
def sline(label, val, bold=False, size=11, header=False):
    ws.append([None, label, val])
    b = ws.cell(row=ws.max_row, column=2); v = ws.cell(row=ws.max_row, column=3)
    b.font = Font(bold=bold or header, size=size, color=(NAVY_FONT if header else TITLE))
    v.font = Font(bold=bold or header, size=size, color=(NAVY_FONT if header else TITLE))
    v.alignment = Alignment(horizontal="center")
    if header:
        for col in (2, 3):
            ws.cell(row=ws.max_row, column=col).fill = PatternFill("solid", fgColor=NAVY)
from collections import Counter
ws.append([None, "Coverage Summary"]); ws.cell(row=ws.max_row, column=2).font = Font(bold=True, size=16)
ws.append([])
sline("Checklist — items by area", "Count", header=True)
for a in CHECK_AREA_ORDER:
    cnt = sum(1 for c in checklist if c.get("area") == a)
    if cnt: sline(a, cnt)
sline("TOTAL", len(checklist), bold=True)
ws.append([])
sline("Components — by category", "Count", header=True)
cc = Counter(c.get("category") for c in components)
for a in COMP_CAT_ORDER:
    if cc.get(a): sline(a, cc[a])
sline("TOTAL", len(components), bold=True)
ws.append([])
sline("Pages — by area", "Count", header=True)
pc = Counter(p.get("area") for p in pages)
for a in PAGE_AREA_ORDER:
    if pc.get(a): sline(a, pc[a])
sline("TOTAL", len(pages), bold=True)

# order: Read Me, Summary, Pages, Components, Checklist
wb.move_sheet("Summary", offset=-3)
wb.save(OUT)
print("Wrote", OUT)
print("Sheets:", wb.sheetnames)
PY = None
